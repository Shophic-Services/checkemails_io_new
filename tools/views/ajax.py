
import csv
import datetime
import pandas as pd
import re, os, uuid
from django.views.generic import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http.response import JsonResponse

from django.http import HttpResponse
from django.utils import timezone
from accounts.models import UserRole
from tools import constants
from tools.models import DataSourceJob, DataSourceItem, EmailData, DataSourceJob, ScraperResult
from tools.tasks.email.ingest import create_datasource_items_task
from django.conf import settings
from subscription.models import ClientCreditSubscription
import openpyxl
import csv
import re
from openpyxl import Workbook, load_workbook
from django.http import HttpResponse

from tools.tasks.scraper.ingest import create_scraper_items_task


class FileUploadView(LoginRequiredMixin, View):
    template_name = 'emailtool/upload_document.html'
    

    def post(self, request, *args, **kwargs):
        """
        Handles chunked CSV uploads for DataSourceJob
        without updating status
        """

        # -----------------------------
        # Validate request data
        # -----------------------------
        uploaded_chunk = request.FILES.get('file')
        original_file_name = request.POST.get('original_file_name')
        updated_file_name = request.POST.get('updated_file_name')
        upload_session_path = request.POST.get('upload_session_path')
        is_last_chunk = int(request.POST.get('is_last_chunk', 0))
        next_chunk_offset = request.POST.get('next_chunk_offset')
        first_row_label = request.POST.get('first_row_label')
        duplicate_confirm = request.POST.get('duplicate_confirm')
        first_row_label = True if first_row_label == 'yes' else False
        duplicate_confirm = True if duplicate_confirm == 'yes' else False
        selected_file_columns = request.POST.get('selected_file_columns', '')
        selected_file_columns = list(map(int, filter(None, selected_file_columns.split(','))))
        source_type = request.POST.get('source', 'bulk')
        source_type = constants.WEBSITE if source_type == 'website' else constants.BULK

        if not uploaded_chunk or not updated_file_name or not original_file_name:
            return JsonResponse({'error': 'Invalid upload request'}, status=400)

        # -----------------------------
        # File path
        # -----------------------------
        upload_dir = os.path.join(settings.MEDIA_ROOT, 'source-reference')
        os.makedirs(upload_dir, exist_ok=True)
        file_path = os.path.join(upload_dir, updated_file_name)
        data_source = None
        # -----------------------------
        # Create or fetch job
        # -----------------------------
        
        try:
            if not upload_session_path or upload_session_path == "null":
                data_source = DataSourceJob.objects.create(
                    name=original_file_name,
                    source_type=source_type,
                    source_reference=f'source-reference/{updated_file_name}',
                    upload_started_at=timezone.now(),
                    first_row_label=first_row_label,
                    duplicate_confirm=duplicate_confirm,
                    selected_file_columns=selected_file_columns,
                    added_by=request.user,
                    belongs_to=request.user,
                    source_format = str(updated_file_name).split('.')[-1].lower()
                )
            else:
                try:
                    data_source = DataSourceJob.objects.get(uuid=upload_session_path)
                except DataSourceJob.DoesNotExist:
                    return JsonResponse({'error': 'Upload session not found'}, status=404)

            # -----------------------------
            # Write chunk to disk
            # -----------------------------
            write_mode = 'ab' if os.path.exists(file_path) else 'wb'
            with open(file_path, write_mode) as destination:
                for chunk in uploaded_chunk.chunks():
                    destination.write(chunk)

            # -----------------------------
            # If last chunk → count emails
            # -----------------------------
            if is_last_chunk:
                total_items = self._count_items(file_path, source_type, data_source.selected_file_columns, data_source.first_row_label, data_source.duplicate_confirm)
                data_source.input_references = total_items
                data_source.total_items = len(total_items)

                active_subscription = ClientCreditSubscription.objects.filter(client=request.user, is_activated=True).order_by("-create_date").first()
                if active_subscription:
                    active_subscription_credit = active_subscription.credit_balance or 0
                    if (active_subscription_credit - active_subscription.credit_reserved) < data_source.total_items: 
                        data_source.delete()
                        return JsonResponse({'error': 'Upload session failed'}, status=500)

                data_source.upload_completed = True
                data_source.upload_completed_at = timezone.now()
                data_source.save()
                return JsonResponse({
                    'message': 'Upload completed successfully',
                })
        except Exception as e:
            print(e)
            if data_source:
                data_source.status = DataSourceJob.UPLOAD_ERROR
                data_source.save()
            return JsonResponse({'error': 'Upload session failed'}, status=500)

        # -----------------------------
        # Continue upload
        # -----------------------------
        return JsonResponse({
            'upload_session_path': str(data_source.uuid),
            'next_chunk_offset': next_chunk_offset
        })
    
    def _count_items(self, file_path, source_type, selected_file_columns=None, first_row_label=False, duplicate_confirm=False):
        """
        Fetch email addresses from CSV or XLSX file.
        `selected_file_columns` can be a list of column indices to read.
        """
        if source_type == constants.WEBSITE:
            TEXT_REGEX = re.compile(
                r'^(https?:\/\/)?(www\.)?[a-z0-9-]+\.[a-z]{2,}(\.[a-z]{2,})?(\/[^\s]*)?$',
                re.IGNORECASE
            )
        else:
            TEXT_REGEX = re.compile(
                r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b'
            )

        total_items = set()
        ext = os.path.splitext(file_path)[1].lower()

        if ext == '.csv':
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as csv_file:
                reader = csv.reader(csv_file)
                if first_row_label:
                    next(reader, None)  # skip header

                for row in reader:
                    # If specific columns are selected, pick only those
                    cells = [row[i] for i in selected_file_columns] if selected_file_columns else row
                    for cell in cells:
                        match = TEXT_REGEX.search(cell.strip())
                        if match:
                            total_items.add(match.group().lower())

        elif ext in ['.xlsx', '.xls']:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active

            rows = ws.iter_rows(values_only=True)
            if first_row_label:
                next(rows, None)  # skip header
            for row in rows:
                cells = [row[i] for i in selected_file_columns] if selected_file_columns else row
                for cell in cells:
                    if cell:
                        match = TEXT_REGEX.search(str(cell).strip())
                        if match:
                            total_items.add(match.group().lower())
        else:
            raise ValueError("Unsupported file type. Please provide CSV or XLSX.")

        return list(set(total_items)) if duplicate_confirm else list(total_items)
    
class ProcessEmailView(LoginRequiredMixin, View):
    
    def post(self, request, *args, **kwargs):
        """
        Endpoint to process emails for a given DataSourceJob
        """
        emails = request.POST.get('emails').split(',')

        wb = Workbook()
        ws = wb.active
        ws.title = "Emails"

        # Add header
        ws.append(["Email Address"])

        # Add emails
        for email in emails:
            ws.append([email.strip()])
        uuid_str = str(uuid.uuid4())
        # File path
        file_path = f"source-reference/{uuid_str}.xlsx"

        # Save file
        
        wb.save(os.path.join(settings.MEDIA_ROOT, file_path))


        sourcename = 'Email Verification - ' + str(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        try:
            data_source = DataSourceJob.objects.create(uuid=uuid_str,name=sourcename, source_reference=file_path, 
                source_type=constants.BULK, 
                upload_started_at=timezone.now(), 
                added_by=request.user, 
                belongs_to=request.user,
                input_references=emails, 
                total_items=len(emails),
                upload_completed = True,
                upload_completed_at = timezone.now(),
                source_format='xlsx', status=DataSourceJob.PENDING)
        except DataSourceJob.DoesNotExist:
            return JsonResponse({'error': 'Upload session not found'}, status=404)
        return JsonResponse({
                    'message': 'Upload completed successfully',
                })

class ProcessWebsiteView(LoginRequiredMixin, View):
    
    def post(self, request, *args, **kwargs):
        """
        Endpoint to process websites for a given DataSourceJob
        """
        websites = request.POST.get('websites').split(',')

        wb = Workbook()
        ws = wb.active
        ws.title = "Websites"

        # Add header
        ws.append(["Website URL"])

        # Add websites
        for website in websites:
            ws.append([website.strip()])
        uuid_str = str(uuid.uuid4())
        # File path
        file_path = f"source-reference/{uuid_str}.xlsx"

        # Save file
        
        wb.save(os.path.join(settings.MEDIA_ROOT, file_path))


        sourcename = 'Website Fetcher - ' + str(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        try:
            data_source = DataSourceJob.objects.create(uuid=uuid_str,name=sourcename, source_reference=file_path, 
                source_type=constants.WEBSITE, 
                upload_started_at=timezone.now(), 
                added_by=request.user, 
                belongs_to=request.user,
                input_references=websites, 
                total_items=len(websites),
                upload_completed = True,
                upload_completed_at = timezone.now(),
                source_format='xlsx', status=DataSourceJob.PENDING)
        except DataSourceJob.DoesNotExist:
            return JsonResponse({'error': 'Upload session not found'}, status=404)
        return JsonResponse({
                    'message': 'Upload completed successfully',
                })

class ValidationInitiateView(LoginRequiredMixin, View):

    
    def post(self, request, *args, **kwargs):
        
        upload_session_path = request.POST.get('upload_session_path')

        try:
            data_source = DataSourceJob.objects.get(uuid=upload_session_path, status=DataSourceJob.PENDING)
        except DataSourceJob.DoesNotExist:
            return JsonResponse({'error': 'Upload session not found'}, status=404)
        
        user_id = request.user.id

        if request.user.is_superuser or request.user.user_role_id in [UserRole.MANAGER, UserRole.TEAM]:
            create_datasource_items_task.delay(data_source.uuid, user_id)
            return JsonResponse({
                'message': 'Started successfully',
                'credit_balance': None,
                'credit_reserved': None,
                'status': 200
            })

        active_subscription = ClientCreditSubscription.objects.filter(client=request.user, is_activated=True).order_by("-create_date").first()
        if active_subscription:
            active_subscription_credit = active_subscription.credit_balance or 0
            if (active_subscription_credit - active_subscription.credit_reserved) >= data_source.total_items:   
                active_subscription.credit_reserved +=  data_source.total_items
                active_subscription.save()
                create_datasource_items_task.delay(data_source.uuid, user_id)
                return JsonResponse({
                    'message': 'Started successfully',
                    'credit_balance': active_subscription_credit,
                    'credit_reserved': active_subscription.credit_reserved,
                    'status': 200
                })
        
        data_source.status = DataSourceJob.ERROR
        data_source.save()
        return JsonResponse({
                    'message': 'Validation Failed',
                    'status': 111
                })
    

class ExtractionInitiateView(LoginRequiredMixin, View):

    
    
    def post(self, request, *args, **kwargs):
        
        upload_session_path = request.POST.get('upload_session_path')

        try:
            data_source = DataSourceJob.objects.get(uuid=upload_session_path, status=DataSourceJob.PENDING)
        except DataSourceJob.DoesNotExist:
            return JsonResponse({'error': 'Upload session not found'}, status=404)
        
        user_id = request.user.id

        if request.user.is_superuser or request.user.user_role_id in [UserRole.MANAGER, UserRole.TEAM]:

            if settings.CELERY_ENABLED:
                create_scraper_items_task.delay(data_source.uuid, user_id)
            else:                
                create_scraper_items_task(data_source.uuid, user_id)
            return JsonResponse({
                'message': 'Started successfully',
                'credit_balance': None,
                'credit_reserved': None,
                'status': 200
            })

        active_subscription = ClientCreditSubscription.objects.filter(client=request.user, is_activated=True).order_by("-create_date").first()
        if active_subscription:
            active_subscription_credit = active_subscription.credit_balance or 0
            if (active_subscription_credit - active_subscription.credit_reserved) >= data_source.total_items:   
                active_subscription.credit_reserved +=  data_source.total_items
                active_subscription.save()

                if settings.CELERY_ENABLED:
                    create_scraper_items_task.delay(data_source.uuid, user_id)
                else:
                    create_scraper_items_task(data_source.uuid, user_id)
                return JsonResponse({
                    'message': 'Started successfully',
                    'credit_balance': active_subscription_credit,
                    'credit_reserved': active_subscription.credit_reserved,
                    'status': 200
                })
        
        data_source.status = DataSourceJob.ERROR
        data_source.save()
        return JsonResponse({
                    'message': 'Validation Failed',
                    'status': 111
                })

class ReportExportView(LoginRequiredMixin, View):

    # --------------------------------------------------
    # Detect email columns using regex
    # --------------------------------------------------
    @staticmethod
    def detect_email_columns(df, sample_rows=50):
        EMAIL_REGEX = re.compile(
            r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b'
        )

        email_columns = []

        for col in df.columns:
            series = (
                df[col]
                .dropna()
                .astype(str)
                .head(sample_rows)
            )

            if series.apply(lambda x: bool(EMAIL_REGEX.search(x))).any():
                email_columns.append(col)

        return email_columns

    # --------------------------------------------------
    # Normalize result_data
    # --------------------------------------------------
    def extract_result_fields(self, result_data: dict) -> dict:
        """
        Final email status logic:
        VALID | INVALID | RISKY | UNKNOWN | ''
        """

        check = result_data.get('check') or {}

        type_code = check.get('type', {}).get('code')
        is_valid = result_data.get('is_valid')
        is_catch_all = check.get('catch_all', {}).get('valid')

        # -------------------------
        # Determine final status
        # -------------------------
        if type_code == EmailData.RISKY:
            final_status = 'RISKY'
        elif is_valid:
            final_status = 'VALID'
        elif type_code == EmailData.UNKNOWN:
            final_status = ''
        else:
            final_status = 'INVALID'

        return {
            'email_status': final_status,
            'catch_all': 'YES' if is_catch_all is True else 'NO' if is_catch_all is False else '',
            'role_based': (
                'YES'
                if check.get('role_based', {}).get('valid') is True
                else 'NO'
                if check.get('role_based', {}).get('valid') is False
                else ''
            ),
            'disposable': (
                'YES'
                if check.get('disposable', {}).get('valid') is True
                else 'NO'
                if check.get('disposable', {}).get('valid') is False
                else ''
            ),
        }

    # --------------------------------------------------
    # POST: Export report
    # --------------------------------------------------
    def post(self, request, *args, **kwargs):
        job_id = request.POST.get('job_id')

        if not job_id:
            return JsonResponse({"error": "job_id required"}, status=400)

        try:
            job = DataSourceJob.objects.get(uuid=job_id)
        except DataSourceJob.DoesNotExist:
            return JsonResponse({'error': 'Upload session not found'}, status=404)

        if not job.source_reference:
            return JsonResponse({"error": "Source file missing"}, status=400)

        # -------------------------
        # Load file
        # -------------------------
        if job.source_format == "csv":
            df = pd.read_csv(job.source_reference)
            ext = "csv"
        else:
            df = pd.read_excel(job.source_reference)
            ext = "xlsx"

        # -------------------------
        # Detect email columns
        # -------------------------
        email_columns = self.detect_email_columns(df)
        if not email_columns:
            return JsonResponse({"error": "Email column not found"}, status=400)

        # Normalize emails
        for col in email_columns:
            df[col] = df[col].astype(str).str.strip().str.lower()

        # Collect unique emails across all columns
        emails = (
            df[email_columns]
            .stack()
            .dropna()
            .unique()
        )

        # -------------------------
        # Fetch results
        # -------------------------
        items = (
            DataSourceItem.objects
            .filter(source_job=job, input_value__in=emails)
            .values('input_value', 'result_data')
        )

        result_map = {
            i['input_value'].lower(): self.extract_result_fields(i['result_data'])
            for i in items
        }

        # -------------------------
        # Insert columns next to each email column
        # -------------------------
        for col in reversed(email_columns):
            col_index = df.columns.get_loc(col)
            new_cols = ['Email Status', 'Catch All', 'Role Based', 'Disposable']
            for i, col_name in enumerate(new_cols):
                df.insert(col_index + 1 + i, f"{col_name}", '')

        # -------------------------
        # Populate values
        # -------------------------
        for col in email_columns:
            for idx, email in df[col].items():
                if not email:
                    continue
                data = result_map.get(email, {})
                df.at[idx, 'Email Status'] = data.get('email_status', '')
                df.at[idx, 'Catch All'] = data.get('catch_all', '')
                df.at[idx, 'Role Based'] = data.get('role_based', '')
                df.at[idx, 'Disposable'] = data.get('disposable', '')

        # -------------------------
        # Download response
        # -------------------------
        filename = f"{job.name}_report.{ext}"

        if ext == "csv":
            response = HttpResponse(content_type="text/csv")
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            df.to_csv(response, index=False)
        else:
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            df.to_excel(response, index=False)

        return response

class ReportExportView(LoginRequiredMixin, View):
    EMAIL_REGEX = re.compile(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b')

    def extract_result_fields(self, result_data: dict) -> dict:
        check = result_data
        type_code = check.get('quality')
        status = check.get('status')
        catch_all = check.get('catch_all')
        if type_code == EmailData.RISKY:
            final_status = 'RISKY'
        elif type_code == EmailData.VALID:
            final_status = 'VALID'
        elif type_code == EmailData.VALID:
        
            final_status = 'INVALID'
        elif type_code == EmailData.UNKNOWN:
            final_status = ''
        else:
            final_status = ''

        return {
            'email_status': final_status,
            'catch_all': 'YES' if catch_all else 'NO' if catch_all is False else '',
            'role_based': 'YES' if check.get('is_role_based') else 'NO',
            'disposable': 'YES' if check.get('status') == EmailData.DISPOSABLE else 'NO' 
        }

    def post(self, request, *args, **kwargs):
        job_id = request.POST.get('job_id')
        if not job_id:
            return JsonResponse({"error": "job_id required"}, status=400)

        try:
            job = DataSourceJob.objects.get(uuid=job_id)
        except DataSourceJob.DoesNotExist:
            return JsonResponse({"error": "Upload session not found"}, status=404)

        if not job.source_reference:
            return JsonResponse({"error": "Source file missing"}, status=400)

        # -------------------------
        # CSV Processing
        # -------------------------
        if job.source_format == "csv":
            with open(job.source_reference.path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                all_rows = list(reader)

            header = all_rows[0]
            email_cols = [i for i, col in enumerate(header) if col and self.EMAIL_REGEX.search(col)]

            if not email_cols:
                # fallback: detect email values in any column
                email_cols = [
                    i for i, col in enumerate(header)
                    if any(self.EMAIL_REGEX.match(str(row[i]).strip()) for row in all_rows[1:])
                ]
            if not email_cols:
                return JsonResponse({"error": "Email column not found"}, status=400)

            # Insert new columns next to each email column safely
            offset = 0
            col_mapping = {}
            for idx in email_cols:
                insert_at = idx + 1 + offset
                for col_name in ['Email Status', 'Catch All', 'Role Based', 'Disposable']:
                    if col_name not in header:
                        header.insert(insert_at, col_name)
                        offset += 1
                        insert_at += 1
                col_mapping[idx] = insert_at - 4  # first inserted column
            all_rows[0] = header

            # Collect unique emails
            emails_set = set()
            for row in all_rows[1:]:
                for idx in email_cols:
                    if idx < len(row) and row[idx]:
                        emails_set.add(row[idx].strip().lower())

            items = DataSourceItem.objects.filter(
                source_job=job, input_value__in=emails_set
            ).values('input_value', 'result_data')

            result_map = {i['input_value'].lower(): self.extract_result_fields(i['result_data']) for i in items}

            # Populate new columns
            for row in all_rows[1:]:
                for email_idx in email_cols:
                    email = row[email_idx].strip().lower() if email_idx < len(row) else ''
                    data = result_map.get(email, {})
                    insert_idx = col_mapping[email_idx]
                    for i, key in enumerate(['email_status', 'catch_all', 'role_based', 'disposable']):
                        row.insert(insert_idx + i, data.get(key, ''))

            # Write CSV response
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{job.name}_report.csv"'
            writer = csv.writer(response, lineterminator='\n')
            writer.writerows(all_rows)
            return response

        # -------------------------
        # XLSX Processing
        # -------------------------
        else:
            wb = load_workbook(job.source_reference.path)
            ws = wb.active
            header = [cell.value for cell in ws[1]]

            email_cols = [i for i, col in enumerate(header) if col and self.EMAIL_REGEX.search(str(col))]
            if not email_cols:
                # fallback: detect emails by values
                email_cols = []
                for i, col in enumerate(header):
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=i+1, max_col=i+1):
                        if row[0].value and self.EMAIL_REGEX.match(str(row[0].value).strip()):
                            email_cols.append(i)
                            break
            if not email_cols:
                return JsonResponse({"error": "Email column not found"}, status=400)

            # Insert new columns next to each email column
            offset = 0
            col_mapping = {}
            for idx in email_cols:
                insert_at = idx + 2 + offset  # openpyxl is 1-indexed
                ws.insert_cols(insert_at, amount=4)
                ws.cell(row=1, column=insert_at, value='Email Status')
                ws.cell(row=1, column=insert_at + 1, value='Catch All')
                ws.cell(row=1, column=insert_at + 2, value='Role Based')
                ws.cell(row=1, column=insert_at + 3, value='Disposable')
                col_mapping[idx] = insert_at
                offset += 4

            # Collect unique emails
            emails_set = set()
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for idx in email_cols:
                    val = row[idx].value
                    if val:
                        emails_set.add(str(val).strip().lower())

            items = DataSourceItem.objects.filter(
                source_job=job, input_value__in=emails_set
            ).values('input_value', 'result_data')

            result_map = {i['input_value'].lower(): self.extract_result_fields(i['result_data']) for i in items}

            # Populate new columns
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
                for email_idx in email_cols:
                    cell_val = row[email_idx].value
                    email = str(cell_val).strip().lower() if cell_val else ''
                    data = result_map.get(email, {})
                    insert_idx = col_mapping[email_idx]
                    ws.cell(row=row_idx + 2, column=insert_idx, value=data.get('email_status', ''))
                    ws.cell(row=row_idx + 2, column=insert_idx + 1, value=data.get('catch_all', ''))
                    ws.cell(row=row_idx + 2, column=insert_idx + 2, value=data.get('role_based', ''))
                    ws.cell(row=row_idx + 2, column=insert_idx + 3, value=data.get('disposable', ''))

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response['Content-Disposition'] = f'attachment; filename="{job.name}_report.xlsx"'
            wb.save(response)
            return response
        
class WebsiteExportView(LoginRequiredMixin, View):

    def clean_for_excel(self, val):
        if isinstance(val, str):
            return val.replace("\x00", "").strip()
        return val


    def join_non_null_unique(self, series):
        return ", ".join(sorted(set(filter(None, series))))
    
    def post(self, request, *args, **kwargs):
        job_id = request.POST.get('job_id')
        if not job_id:
            return JsonResponse({"error": "job_id required"}, status=400)

        try:
            job = DataSourceJob.objects.get(uuid=job_id)
        except DataSourceJob.DoesNotExist:
            return JsonResponse({"error": "Upload session not found"}, status=404)

        results = list(ScraperResult.objects.filter(source_job=job).values('url','final_url','scrape_name','emaildata'))

        

        if not results:
            return HttpResponse("No data available")

        # ✅ Step 1: Normalize data (flatten email list)
        normalized = []
        for r in results:
            emails = r.get("emaildata", []) or [None]

            for email in emails:
                normalized.append({
                    "Website": r.get("url"),
                    "FinalURL": r.get("final_url"),
                    "CompanyName": r.get("scrape_name"),
                    "Email": email,
                    "Status": r.get("status", "Success"),
                    "HTTPStatus": r.get("http_status", 200),
                })

        # ✅ Step 2: Create DataFrame
        df = pd.DataFrame(normalized)

        # ✅ Step 3: Clean data
        df_by_flat = df.apply(lambda col: col.map(self.clean_for_excel))

        # ✅ Step 4: Group emails per site
        df_by_website = (
            df.groupby(
                ["Website", "FinalURL", "Status", "HTTPStatus", "CompanyName"],
                dropna=False,
                as_index=False
            )
            .agg({"Email": self.join_non_null_unique})
        )

        df_by_website = df_by_website.apply(lambda col: col.map(self.clean_for_excel))

        # ✅ Step 5: Write Excel to HTTP response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename="{job.name}_report.xlsx"'

        with pd.ExcelWriter(response, engine="openpyxl") as writer:
            df_by_flat.to_excel(writer, index=False, sheet_name="by_email")
            df_by_website.to_excel(writer, index=False, sheet_name="by_website")
        # wb.save(response)
        return response